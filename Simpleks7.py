###Реализация парсинга из excel###
from openpyxl import load_workbook

wb = load_workbook('C:\XXXX\ИТМО\НИР\Задача№3_Мат_описание\lineranalysis\InputData.xlsx', data_only=True)  # Грузим эксель, первый аргумент - название файла эксель
#Целевая функция
data_for_calculating = []  # Создаем массив для данных
if 'ObjFunction' in wb.sheetnames:  # Проверяем есть ли данный лист в экселе
    ws = wb['ObjFunction']  # Присваиваем переменной ws нужный лист
    for cell in ws[2]:  # Пробегаемся по каждой ячейке во второй строке
        if cell.value == 0:
            break
        # Грузим в массив данные по каждой ячейке
        data_for_calculating.append(cell.value)
else:
    print('List Numbers does not exists')

#Правая сторона ограничений ежемесячного объема производства (тонн)
data_for_calculating1 = []  # Создаем массив для данных
if 'Restrictions' in wb.sheetnames:  # Проверяем есть ли данный лист в экселе
    ws = wb['Restrictions']  # Присваиваем переменной ws нужный лист
    for cell in ws[2]:  # Пробегаемся по каждой ячейке во второй строке
        if cell.value == 0:
            break
        # Грузим в массив данные по каждой ячейке
        data_for_calculating1.append(cell.value)
else:
    print('List Numbers does not exists')


from scipy.optimize import linprog
#linprog решает только задачи минимизации

obj = data_for_calculating[0:7]

#Левая сторона ограничений
lhs_ineq = [[-1, 0, 0, 0, 0, 0, 0],  #-x1: ежемесячный объем производства (тонн)
            [0, -1, 0, 0, 0, 0, 0],  #-x2: ежемесячный объем производства (тонн)
            [0, 0, -1, 0, 0, 0, 0],  #-x3: ежемесячный объем производства (тонн)
            [0, 0, 0, -1, 0, 0, 0],  #-x4: ежемесячный объем производства (тонн)
            [0, 0, 0, 0, -1, 0, 0],  #-x5: ежемесячный объем производства (тонн)
            [0, 0, 0, 0, 0, -1, 0],  #-x6: ежемесячный объем производства (тонн)
            [0, 0, 0, 0, 0, 0, -1],  #-x7: ежемесячный объем производства (тонн)
            [1, 1, 1, 1, 1, 1, 1]]  #x1+x2+x3+x4+x5+x6+x7: ежемесячный объем производства (тонн)

#Правая сторона ограничений ежемесячного объема производства (тонн)
rhs_ineq = data_for_calculating1[0:8]

#obj содержит коэффициенты целевой функции
#lhs_ineq и rhs_ineq содержат коэффициенты из ограничений-неравенств для x.

opt = linprog(c=obj, A_ub=lhs_ineq, b_ub=rhs_ineq, method="simplex")
print(opt)